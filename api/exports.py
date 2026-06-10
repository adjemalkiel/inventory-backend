"""
Générateur de fichiers Excel/CSV (openpyxl) pour les rapports Bâtir Pro.
Chaque fonction reçoit des données déjà calculées et retourne une HttpResponse.
"""

import csv
import io
from decimal import Decimal
from datetime import date

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PRIMARY = "1E3A5F"
ACCENT = "E8F0FE"
WARNING_BG = "FFF3CD"
DANGER_BG = "FFDEDE"
SUCCESS_BG = "D4EDDA"
TOTAL_BG = "F0F4F8"


def _xlsx_response(filename: str) -> tuple[Workbook, HttpResponse]:
    wb = Workbook()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    safe_name = filename.replace(" ", "_")
    response["Content-Disposition"] = f'attachment; filename="{safe_name}"'
    return wb, response


def _csv_response(filename: str) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    safe_name = filename.replace(" ", "_")
    response["Content-Disposition"] = f'attachment; filename="{safe_name}"'
    response.write("\ufeff")  # BOM for Excel compatibility
    return response


def _header_row(ws, columns: list[str], row: int = 1):
    for col_idx, label in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _total_row(ws, row: int, label: str, values: dict[int, Decimal]):
    first_cell = ws.cell(row=row, column=1, value=label)
    first_cell.font = Font(bold=True)
    first_cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
    for col_idx, value in values.items():
        cell = ws.cell(row=row, column=col_idx, value=float(value))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
        cell.number_format = "#,##0.00"
        cell.alignment = Alignment(horizontal="right")


def _autofit_columns(ws, min_width: int = 12, max_width: int = 40):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value or "")) for cell in col if cell.value),
            default=min_width,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _title_block(ws, title: str, subtitle: str = "", org=None):
    row = 1
    if org is not None:
        org_name = org.company_name or "Bâtir Pro"
        ws.cell(row=row, column=1, value=org_name).font = Font(bold=True, size=11, color=PRIMARY)
        row += 1
        if org.company_city:
            ws.cell(row=row, column=1, value=org.company_city).font = Font(italic=True, size=9, color="6B7280")
            row += 1
    ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=13, color=PRIMARY)
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle).font = Font(italic=True, size=9, color="6B7280")


def _finalize_xlsx(wb: Workbook, response: HttpResponse) -> HttpResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    return response


# ─── R1 : Inventaire valorisé ──────────────────────────────

def export_stock_valuation(report: dict, currency: str = "XOF") -> HttpResponse:
    COLS = ["Nom", "SKU", "Catégorie", "Quantité", "Coût unitaire", "Valeur totale"]
    wb, response = _xlsx_response(f"inventaire-valorise-{date.today()}.xlsx")
    ws = wb.active
    ws.title = "Inventaire valorisé"

    _title_block(
        ws,
        f"Inventaire valorisé ({report['method'].upper()})",
        f"Généré le {date.today().strftime('%d/%m/%Y')} — Méthode : {report['method']}",
    )
    _header_row(ws, COLS, row=3)
    ws.freeze_panes = "A4"

    for r_idx, item in enumerate(report["items"], start=4):
        fill = PatternFill("solid", fgColor=ACCENT if r_idx % 2 == 0 else "FFFFFF")
        ws.cell(row=r_idx, column=1, value=item["name"]).fill = fill
        ws.cell(row=r_idx, column=2, value=item["sku"]).fill = fill
        ws.cell(row=r_idx, column=3, value=item.get("category_name") or "—").fill = fill
        c_qty = ws.cell(row=r_idx, column=4, value=float(item["total_quantity"]))
        c_qty.fill = fill
        c_qty.number_format = "#,##0.000"
        c_cu = ws.cell(row=r_idx, column=5, value=float(item["unit_cost"]))
        c_cu.fill = fill
        c_cu.number_format = "#,##0.00"
        c_cu.alignment = Alignment(horizontal="right")
        c_val = ws.cell(row=r_idx, column=6, value=float(item["value"]))
        c_val.fill = fill
        c_val.number_format = "#,##0.00"
        c_val.alignment = Alignment(horizontal="right")

    total_row_idx = len(report["items"]) + 4
    _total_row(ws, total_row_idx, "TOTAL", {6: Decimal(report["grand_total"])})

    _autofit_columns(ws)
    return _finalize_xlsx(wb, response)


# ─── R2 : Coût chantier ────────────────────────────────────

def export_project_cost_breakdown(project, costs: dict, by_category: list) -> HttpResponse:
    wb, response = _xlsx_response(
        f"cout-chantier-{project.reference or project.id}-{date.today()}.xlsx"
    )

    ws1 = wb.active
    ws1.title = "Synthèse"
    _title_block(
        ws1,
        f"Coût chantier — {project.name}",
        f"Ref : {project.reference or '—'} | {date.today().strftime('%d/%m/%Y')}",
    )

    POSTE_MAP = [
        ("Matériaux consommés", "cost_materials"),
        ("Pertes / casses", "cost_losses"),
        ("Main d'œuvre", "cost_labour"),
        ("Sous-traitance", "cost_subcontracting"),
        ("Location équipements", "cost_rental"),
        ("Frais généraux", "cost_overhead"),
    ]
    budget_total = Decimal(str(costs.get("budget_total") or "0"))
    _header_row(ws1, ["Poste", "Coût réalisé", "Budget", "Écart", "%", "Devise"], row=3)
    ws1.freeze_panes = "A4"
    currency = project.currency or "XOF"
    for r_idx, (label, key) in enumerate(POSTE_MAP, start=4):
        ws1.cell(row=r_idx, column=1, value=label)
        actual_val = Decimal(costs.get(key) or "0")
        line_budget = budget_total / Decimal(len(POSTE_MAP)) if budget_total > 0 else Decimal("0")
        ecart = line_budget - actual_val
        pct = round(float(ecart) / float(line_budget) * 100, 1) if line_budget > 0 else None

        c = ws1.cell(row=r_idx, column=2, value=float(actual_val))
        c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
        c = ws1.cell(row=r_idx, column=3, value=float(line_budget))
        c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
        c = ws1.cell(row=r_idx, column=4, value=float(ecart))
        c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
        ws1.cell(row=r_idx, column=5, value=f"{pct:.1f}%" if pct is not None else "—")
        ws1.cell(row=r_idx, column=6, value=currency)
    total_r = len(POSTE_MAP) + 4
    _total_row(ws1, total_r, "COÛT TOTAL", {2: Decimal(str(costs.get("cost_total") or 0))})
    meta_r = total_r + 2
    for label, key in [("Budget", "budget_total"), ("Marge", "margin")]:
        ws1.cell(row=meta_r, column=1, value=label).font = Font(italic=True)
        val = costs.get(key)
        if val is not None:
            c = ws1.cell(row=meta_r, column=2, value=float(Decimal(str(val))))
            c.number_format = "#,##0.00"
        meta_r += 1

    # Sheet 2 — Budget vs Réalisé
    ws2 = wb.create_sheet("Budget vs Réalisé")
    _header_row(ws2, ["Catégorie", "Budget", "Réalisé", "Écart", "%", "Statut"], row=1)
    for r_idx, row in enumerate(by_category, start=2):
        ws2.cell(row=r_idx, column=1, value=row.get("label") or row["category"])
        for col, key in [(2, "budget"), (3, "actual"), (4, "variance")]:
            c = ws2.cell(row=r_idx, column=col, value=float(Decimal(row[key])))
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
        vp = row.get("variance_percent")
        ws2.cell(row=r_idx, column=5, value=f"{vp:.1f}%" if vp is not None else "—")
        statut = "⚠️ DÉPASSÉ" if row.get("over_budget") else "OK"
        c_s = ws2.cell(row=r_idx, column=6, value=statut)
        if row.get("over_budget"):
            c_s.fill = PatternFill("solid", fgColor=WARNING_BG)

    _autofit_columns(ws1)
    _autofit_columns(ws2)
    return _finalize_xlsx(wb, response)


# ─── R3 : Historique mouvements ─────────────────────────────

def export_movements(movements_qs, params: dict) -> HttpResponse:
    COLS = [
        "Date", "Heure", "Référence", "Type", "Article", "SKU",
        "Quantité", "Unité", "Coût unitaire", "Coût total",
        "Source", "Destination", "Chantier", "Statut", "Créé par",
    ]
    date_str = f"{params.get('date_from', '')}_au_{params.get('date_to', '')}"
    wb, response = _xlsx_response(f"mouvements-{date_str}.xlsx")
    ws = wb.active
    ws.title = "Mouvements"
    _header_row(ws, COLS, row=1)
    ws.freeze_panes = "A2"

    for r_idx, mv in enumerate(movements_qs[:10_000], start=2):
        dt = mv.created_at
        ws.cell(row=r_idx, column=1, value=dt.strftime("%d/%m/%Y") if dt else "")
        ws.cell(row=r_idx, column=2, value=dt.strftime("%H:%M") if dt else "")
        ws.cell(row=r_idx, column=3, value=mv.reference_number or "")
        ws.cell(row=r_idx, column=4, value=mv.get_movement_type_display())
        ws.cell(row=r_idx, column=5, value=mv.item.name if mv.item else "")
        ws.cell(row=r_idx, column=6, value=mv.item.sku if mv.item else "")
        c_q = ws.cell(row=r_idx, column=7, value=float(mv.quantity))
        c_q.number_format = "#,##0.000"
        ws.cell(row=r_idx, column=8, value=str(mv.item.unit) if mv.item and mv.item.unit else "")
        for col, val in [(9, mv.unit_price_at_movement), (10, mv.total_cost)]:
            if val is not None:
                c = ws.cell(row=r_idx, column=col, value=float(val))
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
        ws.cell(
            row=r_idx, column=11,
            value=mv.source_storage_location.name if mv.source_storage_location else "",
        )
        ws.cell(
            row=r_idx, column=12,
            value=mv.destination_storage_location.name if mv.destination_storage_location else "",
        )
        ws.cell(row=r_idx, column=13, value=mv.project.name if mv.project else "")
        ws.cell(row=r_idx, column=14, value=mv.get_status_display())
        ws.cell(
            row=r_idx, column=15,
            value=mv.created_by.get_full_name() if mv.created_by else "",
        )

    _autofit_columns(ws)
    return _finalize_xlsx(wb, response)


# ─── R4 : Stock critique (CSV) ──────────────────────────────

def export_critical_stock_csv(items_qs) -> HttpResponse:
    response = _csv_response(f"stock-critique-{date.today()}.csv")
    writer = csv.writer(response)
    writer.writerow([
        "Nom", "SKU", "Catégorie", "Fournisseur",
        "Stock actuel", "Seuil minimum", "Statut", "Coût unitaire",
    ])
    for item in items_qs:
        total_stock = getattr(item, "total_stock", None) or 0
        if float(total_stock) <= 0:
            status_label = "Rupture"
        else:
            status_label = "Stock bas"
        writer.writerow([
            item.name,
            item.sku,
            item.category.name if item.category else "",
            item.supplier_name or "",
            float(total_stock),
            float(item.min_stock) if item.min_stock else 0,
            status_label,
            float(item.unit_price) if item.unit_price else 0,
        ])
    return response


# ─── R5 : Budget vs Réalisé (tous chantiers) ────────────────

def export_all_projects_budget(projects_data: list[dict]) -> HttpResponse:
    COLS = [
        "Chantier", "Référence", "Budget total", "Coût réel",
        "Valeur marché", "Marge", "% marge", "Statut budget",
    ]
    wb, response = _xlsx_response(f"budget-vs-realise-{date.today()}.xlsx")
    ws = wb.active
    ws.title = "Budget vs Réalisé"

    _title_block(ws, "Budget vs Réalisé — Tous chantiers actifs",
                 f"Généré le {date.today().strftime('%d/%m/%Y')}")
    _header_row(ws, COLS, row=3)
    ws.freeze_panes = "A4"

    for r_idx, p in enumerate(projects_data, start=4):
        ws.cell(row=r_idx, column=1, value=p["name"])
        ws.cell(row=r_idx, column=2, value=p.get("reference") or "")
        for col, key in [(3, "budget_total"), (4, "cost_total"), (5, "contract_value"), (6, "margin")]:
            val = p.get(key)
            if val is not None:
                c = ws.cell(row=r_idx, column=col, value=float(Decimal(str(val))))
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
        mp = p.get("margin_percent")
        ws.cell(row=r_idx, column=7, value=f"{mp:.1f}%" if mp is not None else "—")
        over = p.get("over_budget", False)
        c_s = ws.cell(row=r_idx, column=8, value="⚠️ DÉPASSÉ" if over else "OK")
        if over:
            c_s.fill = PatternFill("solid", fgColor=WARNING_BG)

    _autofit_columns(ws)
    return _finalize_xlsx(wb, response)


# ─── R6 : Consommation mensuelle ────────────────────────────

def export_monthly_consumption(data: dict) -> HttpResponse:
    wb, response = _xlsx_response(f"consommation-mensuelle-{data['year']}.xlsx")
    ws = wb.active
    ws.title = "Consommation mensuelle"

    _title_block(ws, f"Consommation mensuelle — {data['year']}",
                 f"Généré le {date.today().strftime('%d/%m/%Y')}")

    # Collect all categories across months
    all_categories = set()
    for month_data in data["months"]:
        for cat_row in month_data["by_category"]:
            all_categories.add(cat_row["category"])
    categories = sorted(all_categories)

    cols = ["Mois"] + categories + ["Total mois"]
    _header_row(ws, cols, row=3)
    ws.freeze_panes = "A4"

    for r_idx, month_data in enumerate(data["months"], start=4):
        ws.cell(row=r_idx, column=1, value=month_data["month_label"])
        cat_costs = {r["category"]: r["total_cost"] for r in month_data["by_category"]}
        for col_idx, cat in enumerate(categories, start=2):
            val = cat_costs.get(cat, "0")
            c = ws.cell(row=r_idx, column=col_idx, value=float(Decimal(val)))
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
        c_t = ws.cell(row=r_idx, column=len(categories) + 2, value=float(Decimal(month_data["month_total"])))
        c_t.number_format = "#,##0.00"
        c_t.font = Font(bold=True)
        c_t.alignment = Alignment(horizontal="right")

    total_row_idx = len(data["months"]) + 4
    _total_row(ws, total_row_idx, "TOTAL ANNUEL", {len(categories) + 2: Decimal(data["grand_total"])})

    _autofit_columns(ws)
    return _finalize_xlsx(wb, response)


# ─── R8 : Performance fournisseurs ──────────────────────────

def export_supplier_performance(data: dict) -> HttpResponse:
    COLS = [
        "Fournisseur", "Nb livraisons", "Quantité totale",
        "Valeur totale", "Prix moyen", "Prix min", "Prix max",
        "Écart prix", "Dernière livraison",
    ]
    wb, response = _xlsx_response(f"performance-fournisseurs-{date.today()}.xlsx")
    ws = wb.active
    ws.title = "Fournisseurs"

    _title_block(ws, "Performance fournisseurs",
                 f"Généré le {date.today().strftime('%d/%m/%Y')}")
    _header_row(ws, COLS, row=3)
    ws.freeze_panes = "A4"

    for r_idx, row in enumerate(data["suppliers"], start=4):
        ws.cell(row=r_idx, column=1, value=row["fournisseur"])
        ws.cell(row=r_idx, column=2, value=row["nb_livraisons"])
        for col, key in [
            (3, "qty_totale"), (4, "valeur_totale"), (5, "prix_moyen"),
            (6, "prix_min"), (7, "prix_max"), (8, "ecart_prix"),
        ]:
            c = ws.cell(row=r_idx, column=col, value=float(Decimal(row[key])))
            c.number_format = "#,##0.00"
            c.alignment = Alignment(horizontal="right")
        ws.cell(row=r_idx, column=9, value=row.get("derniere_livraison") or "")

    _autofit_columns(ws)
    return _finalize_xlsx(wb, response)


# ─── R7 : Transferts inter-sites ────────────────────────────

def export_transfers(movements_qs, params: dict) -> HttpResponse:
    COLS = [
        "Date", "Référence", "Article", "SKU", "Quantité", "Unité",
        "Site source", "Site destination", "Notes", "Créé par",
    ]
    date_str = f"{params.get('date_from', '')}_au_{params.get('date_to', '')}"
    wb, response = _xlsx_response(f"transferts-{date_str}.xlsx")
    ws = wb.active
    ws.title = "Transferts"
    _header_row(ws, COLS, row=1)
    ws.freeze_panes = "A2"

    for r_idx, mv in enumerate(movements_qs[:10_000], start=2):
        dt = mv.created_at
        ws.cell(row=r_idx, column=1, value=dt.strftime("%d/%m/%Y") if dt else "")
        ws.cell(row=r_idx, column=2, value=mv.reference_number or "")
        ws.cell(row=r_idx, column=3, value=mv.item.name if mv.item else "")
        ws.cell(row=r_idx, column=4, value=mv.item.sku if mv.item else "")
        c_q = ws.cell(row=r_idx, column=5, value=float(mv.quantity))
        c_q.number_format = "#,##0.000"
        ws.cell(row=r_idx, column=6, value=str(mv.item.unit) if mv.item and mv.item.unit else "")
        ws.cell(
            row=r_idx, column=7,
            value=mv.source_storage_location.name if mv.source_storage_location else "",
        )
        ws.cell(
            row=r_idx, column=8,
            value=mv.destination_storage_location.name if mv.destination_storage_location else "",
        )
        ws.cell(row=r_idx, column=9, value=mv.comment or "")
        ws.cell(
            row=r_idx, column=10,
            value=mv.created_by.get_full_name() if mv.created_by else "",
        )

    _autofit_columns(ws)
    return _finalize_xlsx(wb, response)
